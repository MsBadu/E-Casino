# The attached file is a list of Charles Dickens’s twitter friends. Write a Python application that loads this data file,
# downloads the user account associated with each user id, the 10 most recent tweets from each user and the collection of
# hashtags used in the tweets. This data needs to be stored in an excel data file.
import urllib
from datetime import datetime
import calendar
import requests
import pprint
import openpyxl
# Copy getEndpoint information from web services work
def getEndpoint(resource, parameters):
    # set the base url
    base_url = 'https://api.twitter.com/'
    # set the api version
    api_version = "2/"
    # set bearer token
    bearer_token = "AAAAAAAAAAAAAAAAAAAAAFzuQQEAAAAAERrEoDqu4kZEZxdZ%2FFlvVwfGDvE%3D1w9la04dvMw0vjCVmcqxyFvos5NWnyzMwdbTJQqzCAdattBPrI"

    # build resource URL
    resource_url = base_url + api_version + resource

    # build headers for authorization
    headers = {
        'Authorization': 'Bearer ' + bearer_token
    }

    # verify resource url
    # print("Getting Endpoint: " + resource_url + "?" + urllib.parse.urlencode(parameters))

    # request data from resource url
    response = requests.get(resource_url, headers=headers, params=parameters)
    # print(response)
    # format response as a python dictionary
    response_data = response.json()
    # return response dictionary to main application
    return response_data

tweetFields = 'attachments,author_id,context_annotations,conversation_id,created_at,entities,geo,id,in_reply_to_user_id,lang,possibly_sensitive,public_metrics,referenced_tweets,reply_settings,source,text,withheld'
userFields = 'created_at,description,entities,id,location,name,pinned_tweet_id,profile_image_url,protected,public_metrics,url,username,verified,withheld'


# Give the friends file a variable name
fileName = "friends.txt"
#Create a pointer to the file defined by file Name
with open (fileName) as filePointer:
    # read in entire file as string into file contents
    fileContents = filePointer.read()
    # Split the file into lines
fileLines = fileContents.split('\n')

#Use openpyxl to create an excel data file (data.xlsx).
testWorkbook = openpyxl.Workbook()
# Create a three worksheet
# Create the header for each one
# The friends worksheet should include screen name, name, location, description, activity (number of tweets),
# friends ratio (number of followers / number of friends), date the account was created
friendsWS = testWorkbook.create_sheet('Friends')
friendsHeader = ['Screen Name', 'Name', 'Location', 'Description', 'Activity', 'Friends Ratio','Account Creation Date']
# Use append to send headers to excel worksheet
friendsWS.append(friendsHeader)

# The tweet data worksheet should have the following data fields: screen name, user id of the tweet author, id string of the tweet,
# character count of each tweet, tweet engagement (number of retweets + favorites), number of mentions, tweet date
tweetsWS = testWorkbook.create_sheet('Tweets')
tweetHeader = ['Screen Name', 'User Id of the Tweet Author', 'Id String of the Tweet',
               'Character Count of each Tweet', 'Tweet Engagements', 'Number of Mentions', 'Tweet Date']
tweetsWS.append(tweetHeader)

# The ‘hashtags’ worksheet should have each hashtag used and the id of the user responsible for using the hashtag.
hastagsWS = testWorkbook.create_sheet('Hashtags')
hastagsHeader = ['Screen Name', 'Hashtags']
hastagsWS.append(hastagsHeader)

# Create a list of starting point for each variable which requires it
location = []
friendRatio = 0
tweetMention = []
hashtagText = []
for fileLine in fileLines:
    username = fileLine
    # Make sure resources is using username
    resources = f'users/by/username/{username}'
    # Have parameters use user_fields
    parameters = {'user.fields': userFields}
    userData = getEndpoint(resources, parameters)
    # pprint.pprint(userData)
    userID = userData['data']['id']
    resources = f'users/{userID}/tweets'
    # Get max 10 parameter from tweeter API references
    parameters = {'tweet.fields': tweetFields,
                  'max_results': 10}
    tweetData = getEndpoint(resources, parameters)
    # pprint.pprint(tweetData)
    screenName = userData['data']['username']
    dickenFriendName = userData['data']['name']
    description = userData['data']['description']
    # Create an if statement whenever you see keyword error
    if 'location' in userData['data']:
        location = userData['data']['location']
    # Have else be a space to ensure the cell is let empty
    else:
        location = ''
    activity = userData['data']['public_metrics']['tweet_count']
    # Create an if statement to ensure that ratio is only found when denominator is 1 or higher
    if userData['data']['public_metrics']['following_count'] > 0:
        friendRatio = (userData['data']['public_metrics']['followers_count']) / (userData['data']['public_metrics']['following_count'])
    else:
        friendRatio = ''
    created_at = userData['data']['created_at']
    # Copy datetime template for web services
    created_at_date = datetime.strptime(created_at,'%Y-%m-%dT%H:%M:%S.%fZ')
    # Create a list with all the variable
    friendLists = [screenName, dickenFriendName, location, description, activity, friendRatio, created_at_date]
    # Use append to send info to excel worksheet
    friendsWS.append(friendLists)
    # pprint.pprint(userData)
    pprint.pprint(tweetData)
    tweetAmount = tweetData['meta']['result_count']
    # One user has no tweets create if statement to check if user tweeted
    if tweetAmount > 0:
        # Create loop to run through each tweet
        for tweet in tweetData['data']:
            screenName = userData['data']['username']
            stringID = tweet['id']
            tweetText = tweet['text']
            tweetEngagment = (tweet['public_metrics']['like_count']) + (tweet['public_metrics']['retweet_count'])
            if 'entities' in tweet and 'mentions' in tweet['entities']:
                tweetMention = tweet['entities']['mentions']
            else:
                tweetMention = ''
                # print (tweetMention)
            tweetCreatedAt = tweet['created_at']
            tweetDate = datetime.strptime(tweetCreatedAt, '%Y-%m-%dT%H:%M:%S.%fZ')
            tweetLists = [screenName, stringID, tweetText, len(tweetText), (tweetEngagment), len(tweetMention),
                          tweetCreatedAt]
            # Use append to send info to excel worksheet
            tweetsWS.append(tweetLists)
            # pprint.pprint(tweet)
            # Place everything under if statement and loop
            if 'entities' in tweet and 'hashtags' in tweet['entities']:
                for tag in tweet['entities']['hashtags']:
                    hashtagText = tag['tag']
                    #print(hashtagText)
                    screenName = userData['data']['username']
                    hastagsList = [screenName, hashtagText]
                    # Use append to send info to excel worksheet
                    hastagsWS.append(hastagsList)

# Save the excel workbook as the end, so you don't get a new workbook everytime
testWorkbook.save('Test3_Workbook.xlsx')
